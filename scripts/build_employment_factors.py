"""Regenerate processed employment-factor outputs and audit workbook."""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from colombia_employment_factors.mappings import (
    CAPEX_OPEX_TECH_MAPPINGS,
    DEFAULT_SOURCE_BY_TECHNOLOGY,
    DEFAULT_SOURCE_NOTES,
    RUTOVITZ_2015_RENEWABLE_LOCAL_MANUFACTURING_SHARE,
    RUTOVITZ_2015_RENEWABLE_LOCAL_MANUFACTURING_TECHS,
    RUTOVITZ_DECLINE_FACTORS,
)
from colombia_employment_factors.projections import (
    add_construction_manufacturing_rows,
    project_employment_factors,
)
from colombia_employment_factors.yearly import (
    build_default_model_employment_factors,
    build_yearly_employment_factors,
)

RAW_INPUT = ROOT / "data" / "raw" / "employment_factors_input.csv"
CAPEX_OPEX_RATIO_INPUT = ROOT / "data" / "audit" / "capex_opex_ratios_2024_2030_2050.csv"
PROCESSED_OUTPUT = ROOT / "data" / "processed" / "employment_factors_with_2030_2050.csv"
YEARLY_OUTPUT = ROOT / "data" / "processed" / "employment_factors_yearly_2024_2055.csv"
MODEL_OUTPUT = ROOT / "data" / "processed" / "employment_factors_model_default_2024_2055.csv"
AUDIT_OUTPUT = ROOT / "data" / "audit" / "employment_learning_curves_2030_2050.xlsx"
PACKAGE_DATA_OUTPUT = (
    ROOT
    / "src"
    / "colombia_employment_factors"
    / "data"
    / "employment_factors_with_2030_2050.csv"
)
PACKAGE_YEARLY_OUTPUT = (
    ROOT
    / "src"
    / "colombia_employment_factors"
    / "data"
    / "employment_factors_yearly_2024_2055.csv"
)
PACKAGE_MODEL_OUTPUT = (
    ROOT
    / "src"
    / "colombia_employment_factors"
    / "data"
    / "employment_factors_model_default_2024_2055.csv"
)


def _rutovitz_mapping_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Technology": technology,
                "Rutovitz_Table9_Tech": values[0],
                "Latin_America_2030_decline_factor": values[1] if values[1] is not None else "-",
                "Applied_multiplier_2015_to_2030": f'=IF(ISNUMBER(C{row_number}),1-C{row_number},"")',
            }
            for row_number, (technology, values) in enumerate(RUTOVITZ_DECLINE_FACTORS.items(), start=2)
        ]
    )


def _catalogue_mapping_table() -> pd.DataFrame:
    rows = []
    for technology, values in CAPEX_OPEX_TECH_MAPPINGS.items():
        mapped_technology, mapping_type, projection_rule = values
        rows.append(
            {
                "Technology": technology,
                "Catalogue_Tech": mapped_technology,
                "Mapping_Type": mapping_type,
                "Projection_Rule": projection_rule,
            }
        )
    return pd.DataFrame(rows)


def _default_source_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Technology": technology,
                "Default_Source": source,
                "Note": DEFAULT_SOURCE_NOTES.get(technology, ""),
            }
            for technology, source in DEFAULT_SOURCE_BY_TECHNOLOGY.items()
        ]
    )


def _add_table_sheet(workbook: Workbook, name: str, data: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(name)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for column_number, column_name in enumerate(data.columns, start=1):
        cell = sheet.cell(1, column_number, column_name)
        cell.fill = header_fill
        cell.font = header_font

    for row_number, row in enumerate(data.itertuples(index=False), start=2):
        for column_number, value in enumerate(row, start=1):
            sheet.cell(row_number, column_number, value)

    reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    table = Table(displayName=name.replace("_", ""), ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"


def _add_overview_sheet(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Overview"
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    headers = ["Workbook", "Purpose", "Key inputs", "Notes", "Source context"]
    for column_number, header in enumerate(headers, start=2):
        cell = sheet.cell(2, column_number, header)
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        [
            "Overview",
            "Navigation and assumptions",
            "User inputs highlighted yellow",
            "Supports ex-post jobs calculations for OSeMOSYS-Colombia",
            "Technology Catalogue, Rutovitz, and employment factor inputs",
        ],
        [
            "Mappings_Rutovitz",
            "Mapping to Rutovitz Table 9",
            "Latin America 2030 decline factors",
            "Construction, manufacturing, and O&M for Rutovitz 2015 rows",
            "Rutovitz 2015 Table 9",
        ],
        [
            "Mappings_Catalogue",
            "Mapping to Colombian catalogue",
            "CAPEX/OPEX ratio technology names",
            "Direct, proxy, partial, or constant mapping rule",
            "Colombian Technology Catalogue",
        ],
        [
            "Method_2030_2050",
            "Transparent formulas",
            "Yellow cells are key inputs",
            "Includes equations and worked inputs",
            "For audit and adjustment",
        ],
        [
            "Employment_Outputs",
            "Final long-format outputs",
            "All original and projected rows",
            "CSV-equivalent dataset",
            "For direct model use",
        ],
        [
            "Yearly_All_Sources",
            "Interpolated yearly outputs",
            "All eligible sources and factor types",
            "Annual values for 2024-2055",
            "Sparse projection table interpolated by source/technology/factor/job/unit",
        ],
        [
            "Default_Source_Mapping",
            "Model default source choices",
            "One source per technology",
            "Used to select the model-ready table",
            "Repository default assumptions",
        ],
        [
            "Model_Default_2024_2055",
            "Model-ready yearly outputs",
            "Default-source rows only",
            "Use this table for OSeMOSYS-style retrieval",
            "Derived from Yearly_All_Sources and Default_Source_Mapping",
        ],
    ]
    for row_number, row in enumerate(rows, start=3):
        for column_number, value in enumerate(row, start=2):
            sheet.cell(row_number, column_number, value)


def _add_method_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Method_2030_2050")
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF59D")

    headers = ["Input type", "Formula", "2030 input", "2050 input", "Example technology", "Comment"]
    for column_number, header in enumerate(headers, start=2):
        cell = sheet.cell(2, column_number, header)
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        [
            "Rutovitz 2015 C/M/O&M",
            "EF_2030 = EF_2015 * (1 - decline_factor_2030)",
            "Latin America Table 9 coefficient",
            "2050/2030 CAPEX or OPEX ratio",
            "Solar PV",
            "2030 uses Rutovitz Table 9; 2050 uses catalogue CAPEX/OPEX transition ratios.",
        ],
        [
            "CAPEX/OPEX ratio",
            "EF_2030 = EF_base * ratio_2030_2024; EF_2050 = EF_2030 * ratio_2050_2030",
            "2030/2024 CAPEX or OPEX ratio",
            "2050/2030 CAPEX or OPEX ratio",
            "Coal power",
            "Used for non-Rutovitz-2015 projection rows mapped to catalogue ratio technologies. For QBIS rows, the 2030/2024 ratio is applied to corrected 2022 base values for simplicity.",
        ],
        [
            "Yearly interpolation",
            "Linear interpolation between known datapoints; EF_2051-2055 = EF_2050",
            "Known points from original/projected rows",
            "No catalogue ratios beyond 2050",
            "All technologies",
            "Rutovitz 2015 values use the 2015 and 2030 points to interpolate 2024-2029.",
        ],
        [
            "Model default source",
            "Filter yearly table by Default_Source_By_Technology",
            "Default source mapping",
            "One source per technology",
            "Offshore wind fixed",
            "QBIS 2023 is used for fixed and floating offshore wind; other source choices are in Default_Source_Mapping.",
        ],
    ]
    for row_number, row in enumerate(rows, start=3):
        for column_number, value in enumerate(row, start=2):
            sheet.cell(row_number, column_number, value)

    for coordinate in ["D3", "E3", "D4", "E4"]:
        sheet[coordinate].fill = input_fill

    sheet["B7"] = "CAPEX/OPEX ratio inputs"
    sheet["B7"].font = Font(name="Calibri", size=11, bold=True)
    ratios = pd.read_csv(CAPEX_OPEX_RATIO_INPUT)
    ratio_inputs = ratios[
        [
            "technology_name",
            "capex_ratio_2030_over_2024",
            "capex_ratio_2050_over_2030",
            "opex_ratio_2030_over_2024",
            "opex_ratio_2050_over_2030",
        ]
    ].rename(
        columns={
            "technology_name": "Catalogue tech",
            "capex_ratio_2030_over_2024": "CAPEX 2030/2024",
            "capex_ratio_2050_over_2030": "CAPEX 2050/2030",
            "opex_ratio_2030_over_2024": "OPEX 2030/2024",
            "opex_ratio_2050_over_2030": "OPEX 2050/2030",
        }
    )
    start_row = 8
    for column_number, column_name in enumerate(ratio_inputs.columns, start=2):
        cell = sheet.cell(start_row, column_number, column_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_number, row in enumerate(ratio_inputs.itertuples(index=False), start=start_row + 1):
        for column_number, value in enumerate(row, start=2):
            sheet.cell(row_number, column_number, value)


def _autofit_columns(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        for column_number in range(1, sheet.max_column + 1):
            width = max(
                len(str(sheet.cell(row_number, column_number).value or ""))
                for row_number in range(1, min(sheet.max_row, 80) + 1)
            )
            sheet.column_dimensions[get_column_letter(column_number)].width = min(max(width + 2, 12), 35)


def _formulaize_projection_rows(sheet, output: pd.DataFrame) -> None:
    columns = {name: index + 1 for index, name in enumerate(output.columns)}
    excel_rows = {index: index + 2 for index in output.index}
    base_rows = {}
    ratio_end_row = 8 + len(pd.read_csv(CAPEX_OPEX_RATIO_INPUT))

    for index, row in output.iterrows():
        key = (
            row["Source"],
            row["Technology"],
            row["Factor_Type"],
            row["Job_Type"],
            row["Unit"],
            row["Year"],
        )
        base_rows[key] = excel_rows[index]

    def value_formula(excel_row: int, base_row: int, ratio_formula: str, column_name: str) -> str:
        column_letter = get_column_letter(columns[column_name])
        return f"={column_letter}{base_row}*{ratio_formula}"

    def capex_opex_ratio_formula(excel_row: int, row: pd.Series) -> str:
        mapped_tech_cell = f"{get_column_letter(columns['Mapped_Catalogue_Tech'])}{excel_row}"
        if not row["Mapped_Catalogue_Tech"]:
            return "1"

        is_opex = row["Factor_Type"] == "O&M"
        if row["Base_Year"] in (2022, 2024) and row["Year"] == 2030:
            ratio_column = "E" if is_opex else "C"
        elif row["Base_Year"] == 2030 and row["Year"] == 2050:
            ratio_column = "F" if is_opex else "D"
        else:
            return "1"

        lookup = (
            f"INDEX(Method_2030_2050!${ratio_column}$9:${ratio_column}${ratio_end_row},"
            f"MATCH({mapped_tech_cell},Method_2030_2050!$B$9:$B${ratio_end_row},0))"
        )
        return f"IFERROR(IF(ISNUMBER({lookup}),{lookup},1),1)"

    def manufacturing_share(row: pd.Series) -> float:
        if (
            row["Source"] == "Rutovitz 2015"
            and row["Technology"] in RUTOVITZ_2015_RENEWABLE_LOCAL_MANUFACTURING_TECHS
        ):
            return RUTOVITZ_2015_RENEWABLE_LOCAL_MANUFACTURING_SHARE
        return 1.0

    for index, row in output.iterrows():
        if (
            row["Source"] == "Rutovitz 2015"
            and row["Year"] == 2030
            and row["Factor_Type"] != "Fuel"
            and row["Method_Applied"] == "Rutovitz Table 9 decline to 2030"
        ):
            key = (
                row["Source"],
                row["Technology"],
                row["Factor_Type"],
                row["Job_Type"],
                row["Unit"],
                row["Base_Year"],
            )
            base_row = base_rows[key]
            excel_row = excel_rows[index]
            factor_column = columns["Rutovitz_2030_Factor"]
            factor_cell = f"{get_column_letter(factor_column)}{excel_row}"
            mapped_tech_cell = f"{get_column_letter(columns['Mapped_Rutovitz_Tech'])}{excel_row}"
            sheet.cell(
                excel_row,
                factor_column,
                f"=INDEX(Mappings_Rutovitz!$D$2:$D$14,MATCH({mapped_tech_cell},Mappings_Rutovitz!$B$2:$B$14,0))",
            )

            for column_name in ("Value", "Value_Numeric", "Projected_Value"):
                sheet.cell(excel_row, columns[column_name], value_formula(excel_row, base_row, factor_cell, column_name))
            continue

        if str(row["Method_Applied"]).startswith(("CAPEX/OPEX ratio projection", "Constant projection")):
            key = (
                row["Source"],
                row["Technology"],
                row["Factor_Type"],
                row["Job_Type"],
                row["Unit"],
                row["Base_Year"],
            )
            base_row = base_rows[key]
            excel_row = excel_rows[index]
            ratio_formula = capex_opex_ratio_formula(excel_row, row)
            for column_name in ("Value", "Value_Numeric", "Projected_Value"):
                sheet.cell(
                    excel_row,
                    columns[column_name],
                    value_formula(excel_row, base_row, ratio_formula, column_name),
                )
            continue

        if row["Method_Applied"] == "Derived Construction&Manufacturing":
            excel_row = excel_rows[index]
            construction_key = (
                row["Source"],
                row["Technology"],
                "Construction",
                row["Job_Type"],
                row["Unit"],
                row["Year"],
            )
            manufacturing_key = (
                row["Source"],
                row["Technology"],
                "Manufacturing",
                row["Job_Type"],
                row["Unit"],
                row["Year"],
            )
            if construction_key not in base_rows or manufacturing_key not in base_rows:
                continue
            construction_row = base_rows[construction_key]
            manufacturing_row = base_rows[manufacturing_key]
            share = manufacturing_share(row)
            for column_name in ("Value", "Value_Numeric", "Projected_Value"):
                column_letter = get_column_letter(columns[column_name])
                sheet.cell(
                    excel_row,
                    columns[column_name],
                    f"={column_letter}{construction_row}+{share}*{column_letter}{manufacturing_row}",
                )


def write_audit_workbook(
    output: pd.DataFrame,
    yearly_output: pd.DataFrame,
    model_output: pd.DataFrame,
    path: Path,
) -> None:
    workbook = Workbook()
    _add_overview_sheet(workbook)
    _add_table_sheet(workbook, "Mappings_Rutovitz", _rutovitz_mapping_table())
    _add_table_sheet(workbook, "Mappings_Catalogue", _catalogue_mapping_table())
    _add_method_sheet(workbook)
    _add_table_sheet(workbook, "Employment_Outputs", output)
    _formulaize_projection_rows(workbook["Employment_Outputs"], output)
    _add_table_sheet(workbook, "Yearly_All_Sources", yearly_output)
    _add_table_sheet(workbook, "Default_Source_Mapping", _default_source_table())
    _add_table_sheet(workbook, "Model_Default_2024_2055", model_output)
    _autofit_columns(workbook)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    source = pd.read_csv(RAW_INPUT)
    capex_opex_ratios = pd.read_csv(CAPEX_OPEX_RATIO_INPUT)
    output = add_construction_manufacturing_rows(project_employment_factors(source, capex_opex_ratios))
    yearly_output = build_yearly_employment_factors(output)
    model_output = build_default_model_employment_factors(yearly_output)

    PROCESSED_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    output.to_csv(PROCESSED_OUTPUT, index=False, float_format="%.12g")
    yearly_output.to_csv(YEARLY_OUTPUT, index=False, float_format="%.12g")
    model_output.to_csv(MODEL_OUTPUT, index=False, float_format="%.12g")

    PACKAGE_DATA_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(PROCESSED_OUTPUT, PACKAGE_DATA_OUTPUT)
    shutil.copyfile(YEARLY_OUTPUT, PACKAGE_YEARLY_OUTPUT)
    shutil.copyfile(MODEL_OUTPUT, PACKAGE_MODEL_OUTPUT)

    write_audit_workbook(output, yearly_output, model_output, AUDIT_OUTPUT)
    print(
        {
            "csv": str(PROCESSED_OUTPUT),
            "yearly_csv": str(YEARLY_OUTPUT),
            "model_csv": str(MODEL_OUTPUT),
            "xlsx": str(AUDIT_OUTPUT),
            "rows": len(output),
            "yearly_rows": len(yearly_output),
            "model_rows": len(model_output),
            "technologies": output["Technology"].nunique(),
            "model_technologies": model_output["Technology"].nunique(),
        }
    )


if __name__ == "__main__":
    main()
